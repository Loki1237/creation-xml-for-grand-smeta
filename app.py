import tkinter as tk
from tkinter import StringVar, filedialog, messagebox
import openpyxl
import re as reg_exp
import xml.etree.ElementTree as ET
from xml.dom import minidom
from xml.etree.ElementTree import Element, SubElement

class App:
    excel_file = None
    xml_file_name = None

    def init_window(self):
        window = tk.Tk()
        window.title('XML signatures creator')
        window.configure(padx=20, pady=20)
        window.rowconfigure([2], pad=30)

        self.xml_file_name = StringVar(master=window)

        self.lbl_file_pick = tk.Label(text='Файл excel')
        self.lbl_file_pick.grid(row=0, column=0, sticky='w')

        self.btn_file_pick_dialog_show= tk.Button(text='Выбрать', command=self.show_file_pick_dialog)
        self.btn_file_pick_dialog_show.grid(row=0, column=1, sticky='w')

        self.lbl_excel_file_name = tk.Label(text='')
        self.lbl_excel_file_name.grid(row=1, column=0, sticky='w', columnspan=2)

        self.lbl_xml_file_name = tk.Label(text='Имя файла xml')
        self.lbl_xml_file_name.grid(row=2, column=0, sticky='w')

        self.ent_xml_file_name_input = tk.Entry(textvariable=self.xml_file_name, width=40)
        self.ent_xml_file_name_input.grid(row=2, column=1)
        self.ent_xml_file_name_input.insert(0, 'signatures')

        self.btn_save = tk.Button(text='Создать', command=self.generate_xml)
        self.btn_save.grid(row=3, column=1, sticky='e')

        window.mainloop()

    def show_file_pick_dialog(self):
        file_pick_dialog = filedialog.Open(filetypes=[('Файлы excel', '*.xlsx')])
        excel_file = file_pick_dialog.show()

        if excel_file:
            self.excel_file = excel_file
            self.lbl_excel_file_name.config(text=self.excel_file or '')

    def generate_xml(self):
        if self.excel_file:
            xml_name = self.xml_file_name.get()

            if xml_name:
                xlsx = openpyxl.load_workbook(self.excel_file)
                sheet = xlsx.active

                GrandSmeta = Element('GrandSmeta')
                OrgRoot = SubElement(GrandSmeta, 'OrgRoot')
                Item = SubElement(OrgRoot, 'Item', { 'Caption': 'МГК Гранд' })
                Attributes = SubElement(Item, 'Attributes')

                SubElement(Attributes, 'Item', { 'Caption': 'Адрес', 'ID': '800', 'Value': 'г.Москва, ул.Страстной бульвар, 4/3, стр.3, оф.99' })
                SubElement(Attributes, 'Item', { 'Caption': 'Телефон', 'ID': '810', 'Value': '(495) 105-77-88' })
                SubElement(Attributes, 'Item', { 'Caption': 'Факс', 'ID': '820', 'Value': '(495) 105-77-88' })
                SubElement(OrgRoot, 'Item', { 'Caption': 'Ваше представительство МГК Гранд' })

                Group = SubElement(OrgRoot, 'Group', { 'Caption': 'ИНН' })

                for row_number in range(sheet.max_row):
                    name = sheet.cell(row = row_number + 1, column = 1).value
                    tax_id = sheet.cell(row = row_number + 1, column = 2).value

                    Item = SubElement(Group, 'Item', { 'Caption': str(name) })
                    Attributes = SubElement(Item, 'Attributes')
                    SubElement(Attributes, 'Item', { 'Caption': 'ИНН', 'ID': '830', 'Value': str(tax_id) })

                with open('./' + xml_name + '.xml', 'w') as f:
                    xml = ET.tostring(GrandSmeta, encoding='windows-1251')
                    xml = xml.decode(encoding='windows-1251')
                    reparsed = minidom.parseString(xml)
                    pretty_xml = reparsed.toprettyxml(indent='  ')

                    re = reg_exp.compile('\n')
                    pretty_xml_string = '<?xml version="1.0" encoding="windows-1251"?>\n' + re.split(pretty_xml, maxsplit=1)[1]

                    f.write(pretty_xml_string)

            else:
                messagebox.showinfo('Ошибка', 'Введите имя создаваемого файла xml')
                return
        
        else:
            messagebox.showinfo('Ошибка', 'Выберите файл excel')
            return

app = App()
app.init_window()
